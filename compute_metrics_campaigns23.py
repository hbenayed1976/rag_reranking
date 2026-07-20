"""
compute_metrics_campaigns23.py — Agrège les métriques des Campagnes 2 et 3
une fois les fichiers xlsx annotés (cf. build_annotation_batches_campaigns23.py),
et calcule un test de Wilcoxon signé apparié (Scénario A vs Scénario B, par
question) pour les 2 champs numériques.

Champs traités :
  - retrieval_relevance_1_5    -> moyenne ± écart-type, Wilcoxon
  - retrieval_sufficiency_0_2  -> taux % score=2, Wilcoxon
  - primary_error_source       -> distribution (catégoriel, parmi les
                                   réponses incorrectes) ; PAS de test de
                                   Wilcoxon dessus (non ordinal/non
                                   apparié au sens numérique) ; Cohen's κ
                                   catégoriel calculé séparément par
                                   scénario si possible.

Les valeurs numériques remontées par question sont la MOYENNE des deux
annotateurs (faute d'adjudication formelle) ; le test de Wilcoxon est
calculé sur ces moyennes par question, apparié Scénario A vs Scénario B.

Usage:
    python compute_metrics_campaigns23.py --campaign 2
    python compute_metrics_campaigns23.py --campaign 3
    python compute_metrics_campaigns23.py --campaign all
"""

import os
import glob
import argparse
import statistics as stats
from collections import defaultdict

from openpyxl import load_workbook

try:
    from scipy.stats import wilcoxon
    HAS_SCIPY = True
except ImportError:
    HAS_SCIPY = False

try:
    from sklearn.metrics import cohen_kappa_score
    HAS_SKLEARN = True
except ImportError:
    HAS_SKLEARN = False

import config

ANNOTATION_RESULTS_DIR = os.path.join(config.RESULTS_DIR, "annotation")

NUMERIC_CRITERIA = ["retrieval_relevance_1_5", "retrieval_sufficiency_0_2"]
NOT_APPLICABLE_LABEL = "N/A (correct answer)"


def _read_xlsx_rows(path: str) -> list:
    wb = load_workbook(path, read_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter, None)
    if header is None:
        wb.close()
        return []
    rows = [dict(zip(header, row)) for row in rows_iter]
    wb.close()
    return rows


def _numeric_or_none(value):
    if value is None:
        return None
    s = str(value).strip()
    if s == "":
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _is_correct(row: dict) -> bool:
    return str(row.get("is_correct")).strip().lower() == "true"


def load_files(campaign_key: str) -> dict:
    """Renvoie {scenario_type ('A'/'B'): {annotator: [rows...]}} pour une campagne."""
    pattern = os.path.join(ANNOTATION_RESULTS_DIR, f"annotation_campaign{campaign_key}_batch*.xlsx")
    files = sorted(glob.glob(pattern))
    data = defaultdict(lambda: defaultdict(list))
    for path in files:
        rows = _read_xlsx_rows(path)
        for row in rows:
            scenario_type = row.get("scenario_type")
            annotator = row.get("annotator")
            if scenario_type is None or annotator is None:
                continue
            data[scenario_type][annotator].append(row)
    return data


def per_question_means(rows_by_annotator: dict, crit: str) -> dict:
    """Renvoie {question_id: valeur moyennée sur les annotateurs} pour un critère numérique."""
    by_qid = defaultdict(list)
    for annotator, rows in rows_by_annotator.items():
        for row in rows:
            v = _numeric_or_none(row.get(crit))
            if v is not None:
                by_qid[row["question_id"]].append(v)
    return {qid: stats.mean(vals) for qid, vals in by_qid.items() if vals}


def summarize_scenario(rows_by_annotator: dict) -> dict:
    summary = {"n_rows_per_annotator": {a: len(r) for a, r in rows_by_annotator.items()}}

    for crit in NUMERIC_CRITERIA:
        means = list(per_question_means(rows_by_annotator, crit).values())
        if not means:
            summary[f"{crit}_mean"] = None
            summary[f"{crit}_std"] = None
        else:
            summary[f"{crit}_mean"] = round(stats.mean(means), 2)
            summary[f"{crit}_std"] = round(stats.pstdev(means), 2) if len(means) > 1 else 0.0

    suf_means = list(per_question_means(rows_by_annotator, "retrieval_sufficiency_0_2").values())
    if suf_means:
        summary["retrieval_sufficiency_rate_pct"] = round(
            100 * sum(1 for v in suf_means if v >= 1.5) / len(suf_means), 1
        )
    else:
        summary["retrieval_sufficiency_rate_pct"] = None

    # --- Primary Error Source : distribution parmi les réponses incorrectes ---
    error_source_counts = defaultdict(int)
    n_incorrect_rows = 0
    for annotator, rows in rows_by_annotator.items():
        for row in rows:
            if _is_correct(row):
                continue
            source = (row.get("primary_error_source") or "").strip()
            if not source or source.upper().startswith("N/A"):
                continue
            n_incorrect_rows += 1
            error_source_counts[source] += 1
    summary["primary_error_source_distribution"] = {
        source: {
            "count": count,
            "pct": round(100 * count / n_incorrect_rows, 1) if n_incorrect_rows else None,
        }
        for source, count in sorted(error_source_counts.items())
    }
    summary["n_incorrect_rows_with_error_source"] = n_incorrect_rows

    return summary


def cohen_kappa_error_source(rows_by_annotator: dict) -> float:
    """Cohen's kappa catégoriel sur primary_error_source, entre les 2
    annotateurs, calculé sur les questions incorrectes où les deux ont
    rempli une vraie catégorie (pas N/A)."""
    annotators = sorted(rows_by_annotator.keys())
    if len(annotators) != 2 or not HAS_SKLEARN:
        return None
    a1, a2 = annotators
    by_qid = {a: {r["question_id"]: r for r in rows_by_annotator[a]} for a in annotators}
    common_qids = set(by_qid[a1].keys()) & set(by_qid[a2].keys())

    pairs = []
    for qid in common_qids:
        r1, r2 = by_qid[a1][qid], by_qid[a2][qid]
        if _is_correct(r1) or _is_correct(r2):
            continue
        s1 = (r1.get("primary_error_source") or "").strip()
        s2 = (r2.get("primary_error_source") or "").strip()
        if s1 and s2 and not s1.upper().startswith("N/A") and not s2.upper().startswith("N/A"):
            pairs.append((s1, s2))
    if len(pairs) < 2:
        return None
    try:
        return round(cohen_kappa_score([p[0] for p in pairs], [p[1] for p in pairs]), 3)
    except ValueError:
        return None


def wilcoxon_a_vs_b(data: dict) -> dict:
    """Test de Wilcoxon signé apparié (Scénario A vs B), sur les 2 champs
    numériques uniquement (Primary Error Source est catégoriel, exclu)."""
    results = {}
    if "A" not in data or "B" not in data:
        return {crit: None for crit in NUMERIC_CRITERIA}
    if not HAS_SCIPY:
        print("⚠️ scipy indisponible : le test de Wilcoxon sera sauté (installez scipy).")
        return {crit: None for crit in NUMERIC_CRITERIA}

    for crit in NUMERIC_CRITERIA:
        means_a = per_question_means(data["A"], crit)
        means_b = per_question_means(data["B"], crit)
        common_qids = sorted(set(means_a.keys()) & set(means_b.keys()))
        x = [means_a[qid] for qid in common_qids]
        y = [means_b[qid] for qid in common_qids]
        diffs = [a - b for a, b in zip(x, y)]
        if len(diffs) < 2 or all(d == 0 for d in diffs):
            results[crit] = {"n": len(diffs), "statistic": None, "p_value": None}
            continue
        try:
            stat, p = wilcoxon(x, y)
            results[crit] = {"n": len(diffs), "statistic": round(float(stat), 3), "p_value": round(float(p), 6)}
        except ValueError as e:
            results[crit] = {"n": len(diffs), "statistic": None, "p_value": None, "error": str(e)}
    return results


def print_report(campaign_key: str, summaries: dict, wilcoxon_results: dict, kappas: dict):
    spec_label = {
        "2": "Campagne 2 — Fanar-9B, CAMeLBERT — mini_reranker (A) vs cross_encoder (B)",
        "3": "Campagne 3 — Llama3-8B, CAMeLBERT — dual_mini_ara (A) vs dual_mini_cross (B)",
    }.get(campaign_key, f"Campagne {campaign_key}")

    print(f"\n{'=' * 70}")
    print(spec_label)
    print(f"{'=' * 70}")

    for scenario_type in ("A", "B"):
        s = summaries.get(scenario_type, {})
        print(f"\n  Scénario {scenario_type} : {s.get('n_rows_per_annotator')}")
        print(f"    Retrieval Relevance   : {s.get('retrieval_relevance_1_5_mean')} ± {s.get('retrieval_relevance_1_5_std')}")
        print(f"    Retrieval Sufficiency : {s.get('retrieval_sufficiency_rate_pct')}%")
        dist = s.get("primary_error_source_distribution", {})
        n_err = s.get("n_incorrect_rows_with_error_source", 0)
        print(f"    Primary Error Source (n={n_err} annotations sur réponses incorrectes) :")
        for source, d in dist.items():
            print(f"        - {source}: {d['count']} ({d['pct']}%)")
        k = kappas.get(scenario_type)
        print(f"    Cohen's κ (Primary Error Source) : {k}")

    print(f"\n  Test de Wilcoxon signé apparié (A vs B) :")
    for crit, res in wilcoxon_results.items():
        if res is None:
            print(f"    - {crit}: N/A")
        else:
            print(f"    - {crit}: n={res['n']}, W={res.get('statistic')}, p={res.get('p_value')}")


def write_csv(campaign_key: str, summaries: dict, out_path: str):
    import csv
    dirpath = os.path.dirname(out_path)
    if dirpath:
        os.makedirs(dirpath, exist_ok=True)
    fieldnames = ["campaign", "scenario_type",
                  "retrieval_relevance_1_5_mean", "retrieval_relevance_1_5_std",
                  "retrieval_sufficiency_rate_pct",
                  "n_incorrect_rows_with_error_source",
                  "primary_error_source_distribution"]
    write_header = not os.path.exists(out_path)
    with open(out_path, "a", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        if write_header:
            writer.writeheader()
        for scenario_type, s in summaries.items():
            row = {"campaign": campaign_key, "scenario_type": scenario_type}
            row.update({k: s.get(k) for k in fieldnames if k not in ("campaign", "scenario_type")})
            writer.writerow(row)
    print(f"\n💾 Résumé ajouté à {out_path}")


def main():
    parser = argparse.ArgumentParser(
        description="Agrège les métriques des Campagnes 2/3 (Retrieval Relevance, Retrieval "
                     "Sufficiency, Primary Error Source) et calcule un test de Wilcoxon signé "
                     "apparié Scénario A vs B sur les 2 champs numériques."
    )
    parser.add_argument("--campaign", required=True, choices=["2", "3", "all"])
    parser.add_argument("--output", default=os.path.join(ANNOTATION_RESULTS_DIR, "metrics_summary_campaigns23.csv"))
    args = parser.parse_args()

    if not HAS_SCIPY:
        print("⚠️ scipy indisponible : installez-le pour le test de Wilcoxon (pip install scipy --break-system-packages).")

    campaigns_to_run = ["2", "3"] if args.campaign == "all" else [args.campaign]

    for campaign_key in campaigns_to_run:
        data = load_files(campaign_key)
        if not data:
            print(f"❌ Aucun fichier annotation_campaign{campaign_key}_batch*.xlsx trouvé dans "
                  f"{ANNOTATION_RESULTS_DIR}/")
            continue

        summaries = {scenario_type: summarize_scenario(rows_by_annotator)
                     for scenario_type, rows_by_annotator in data.items()}
        kappas = {scenario_type: cohen_kappa_error_source(rows_by_annotator)
                  for scenario_type, rows_by_annotator in data.items()}
        wilcoxon_results = wilcoxon_a_vs_b(data)

        print_report(campaign_key, summaries, wilcoxon_results, kappas)
        write_csv(campaign_key, summaries, args.output)


if __name__ == "__main__":
    main()
