"""
build_annotation_batches_campaigns23.py — Construit les 4 lots d'annotation
(2 par campagne) pour compléter le Protocole A, en plus du Protocole B (Baseline vs Best RAG, déjà réalisé séparément) :

  Campagne 2 — Validation des re-rankers arabes vs multilingues
      LLM fixé      : Fanar-9B
      Embedding fixé: CAMeLBERT (meilleur embedding global de l'étude,
                      Table 10 ; apparaît dans les meilleures configs des
                      3 LLM, choix méthodologiquement solide pour isoler
                      l'effet du re-ranking sans variabilité d'embedding)
      Scénario A    : mini_reranker
      Scénario B    : cross_encoder
      Justification : Table 12, différences significatives (Holm) pour
                       Fanar-9B où cross_encoder est systématiquement la
                       config perdante (mini_reranker > cross_encoder,
                       p_holm=0.0029 ; ara_reranker > cross_encoder,
                       p_holm=0.0099 ; dual_mini_ara > cross_encoder,
                       p_holm=0.0069).

  Campagne 3 — Validation de l'effet du deuxième étage de re-ranking
      LLM fixé      : Llama3-8B
      Embedding fixé: CAMeLBERT
      Scénario A    : dual_mini_ara
      Scénario B    : dual_mini_cross
      Justification : Table 12, l'effet le plus significatif de toute
                       l'étude (dual_mini_ara > dual_mini_cross,
                       p_holm=0.000014 — plus significatif encore que
                       dual_mini_ara vs cross_encoder, p_holm=0.000225).

CHAMPS D'ANNOTATION (3, alignés sur les conventions actées pour le
Protocole B — Faithfulness et Hallucination retirés pour homogénéité) :
  - retrieval_relevance_1_5      : 1-5, passages top-5 topiquement pertinents ?
  - retrieval_sufficiency_0_2    : 0-2 — les passages suffisent-ils à
                                    répondre correctement (indépendamment
                                    de ce que le LLM a généré) ?
  - primary_error_source         : uniquement si is_correct = False. Les
                                    Scénarios A et B des Campagnes 2/3
                                    impliquent TOUS DEUX du retrieval (pas
                                    de condition Baseline ici, contrairement
                                    au Protocole B) : un seul jeu de
                                    catégories est donc utilisé, celui du
                                    Protocole B côté RAG : Retrieval failure
                                    / Generator failure / Both / Ambiguous
                                    question.
  + annotator_notes (texte libre)

ECHANTILLONNAGE : 40 questions par campagne (20 correctes / 20 incorrectes,
tirage FIXE — pas la règle "toutes les erreurs + complément" utilisée pour
les Protocoles A/B), LES MÊMES questions réutilisées pour le Scénario A et
le Scénario B de la campagne (comparaison appariée).

Prérequis :
  - results/generation/rag_{llm}.csv doit contenir CAMeLBERT + les 2
    scénarios de la campagne, sur les 140 questions.
  - Les caches de contexte (build_context_cache.py) doivent exister pour
    CAMeLBERT + mini_reranker, CAMeLBERT + cross_encoder (Campagne 2) et
    CAMeLBERT + dual_mini_ara, CAMeLBERT + dual_mini_cross (Campagne 3).

Usage:
    python build_annotation_batches_campaigns23.py --campaign 2
    python build_annotation_batches_campaigns23.py --campaign 3
    python build_annotation_batches_campaigns23.py --campaign all
"""

import os
import json
import time
import random
import argparse

from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation

import config
from pipeline_utils import (
    EmbeddingRetriever,
    HFGenerator,
    load_all_rerankers,
    unload_rerankers,
    print_environment_banner,
)
from answer_extraction_fixed import validate_answer, extract_answer_letter

from build_annotation_batches import (
    retrieve_with_logging,
    load_completed_keys_xlsx,
    append_row_xlsx,
    load_all_results_deduped,
    ANNOTATORS,
    ANNOTATION_RESULTS_DIR,
)

FIXED_EMBEDDING = "CAMeLBERT"

CAMPAIGN_SPECS = {
    "2": {
        "name": "campaign2_arabic_vs_multilingual_reranker",
        "llm": "Fanar-9B",
        "embedding": FIXED_EMBEDDING,
        "scenario_a": "mini_reranker",
        "scenario_b": "cross_encoder",
        "label": "Arabic re-rankers vs multilingual cross-encoder",
    },
    "3": {
        "name": "campaign3_arabic_vs_multilingual_second_stage",
        "llm": "Llama3-8B",
        "embedding": FIXED_EMBEDDING,
        "scenario_a": "dual_mini_ara",
        "scenario_b": "dual_mini_cross",
        "label": "Arabic vs multilingual second-stage re-ranking",
    },
}

# Primary Error Source : un seul jeu de catégories, car Scénarios A et B
# impliquent tous deux du retrieval dans les Campagnes 2/3 (contrairement
# au Protocole B où S1/Baseline n'a pas de retrieval). Catégories
# reprises telles quelles du Protocole B côté RAG, pour homogénéité.
ERROR_SOURCE_OPTIONS = ["Retrieval failure", "Generator failure", "Both", "Ambiguous question"]
NOT_APPLICABLE_LABEL = "N/A (correct answer)"

# Base du tirage 20/20 (cf. docstring) : "B" = scénario B, "A" = scénario A.
SAMPLING_BASIS = "B"

N_CORRECT = 20
N_INCORRECT = 20
SEED_DEFAULT = 42


def rag_results_path(llm_label: str) -> str:
    return os.path.join(config.GENERATION_RESULTS_DIR, f"rag_{llm_label}.csv")


def campaign_batch_path(campaign_key: str, scenario_type: str, llm_label: str,
                         scenario_name: str, annotator: str) -> str:
    scenario_slug = scenario_name.replace(" ", "_").lower()
    fname = (f"annotation_campaign{campaign_key}_batch{scenario_type}_{llm_label}_"
             f"{FIXED_EMBEDDING}_{scenario_slug}_{annotator}.xlsx")
    return os.path.join(ANNOTATION_RESULTS_DIR, fname)


OUTPUT_KEY_COLS = ["llm", "scenario_type", "question_id"]


def sample_fixed_20_20(question_ids: list, is_correct_map: dict, seed: int,
                        n_correct: int = N_CORRECT, n_incorrect: int = N_INCORRECT) -> list:
    """Tirage 20 correctes / 20 incorrectes (au lieu de la règle 'toutes
    les erreurs + complément' des Protocoles A/B), avec COMPENSATION
    croisée si une catégorie est en déficit : si moins de n_incorrect
    incorrectes sont disponibles, le manque est comblé par des correctes
    supplémentaires (et inversement), de façon à toujours viser un total
    de n_correct + n_incorrect questions, tant que le total de questions
    disponibles le permet."""
    rng = random.Random(seed)
    incorrect = [qid for qid in question_ids if not is_correct_map[qid]]
    correct = [qid for qid in question_ids if is_correct_map[qid]]
    n_total_target = n_correct + n_incorrect

    n_take_incorrect = min(len(incorrect), n_incorrect)
    n_take_correct = min(len(correct), n_correct)

    deficit_incorrect = n_incorrect - n_take_incorrect
    deficit_correct = n_correct - n_take_correct

    if deficit_incorrect > 0:
        print(f"   ⚠️ Seulement {len(incorrect)} cas incorrects disponibles (< {n_incorrect} demandés) : "
              f"tous inclus, {deficit_incorrect} complétés par des correctes supplémentaires si possible.")
        n_take_correct = min(len(correct), n_take_correct + deficit_incorrect)

    if deficit_correct > 0:
        print(f"   ⚠️ Seulement {len(correct)} cas corrects disponibles (< {n_correct} demandés) : "
              f"tous inclus, {deficit_correct} complétés par des incorrectes supplémentaires si possible.")
        n_take_incorrect = min(len(incorrect), n_take_incorrect + deficit_correct)

    sample_incorrect = (rng.sample(incorrect, n_take_incorrect)
                         if n_take_incorrect < len(incorrect) else list(incorrect))
    sample_correct = (rng.sample(correct, n_take_correct)
                       if n_take_correct < len(correct) else list(correct))

    sample = sample_incorrect + sample_correct
    if len(sample) < n_total_target:
        print(f"   ⚠️ Seulement {len(sample)}/{n_total_target} questions au total disponibles "
              f"({len(incorrect)} incorrectes + {len(correct)} correctes) : impossible d'atteindre "
              f"{n_total_target}, échantillon complet inclus.")

    rng.shuffle(sample)
    return sample


def build_row(campaign_key, campaign_label, scenario_type, scenario_name,
              llm_label, embedding_name, qid, q, options_text, correct_letter,
              passages, generated_answer, is_correct, predicted_letter,
              annotator, error_message=None):
    row = {
        "campaign": campaign_key,
        "campaign_label": campaign_label,
        "scenario_type": scenario_type,     # "A" ou "B"
        "scenario_name": scenario_name,     # ex. "mini_reranker" / "cross_encoder"
        "llm": llm_label,
        "embedding": embedding_name,
        "question_id": qid,
        "annotator": annotator,
        "question": q,
        "options": options_text,
        "correct_answer_letter": correct_letter,
    }

    for i in range(1, config.FINAL_N + 1):
        if i <= len(passages):
            rank, score, text = passages[i - 1]
        else:
            rank, score, text = "", "", ""
        row[f"passage_{i}_rank"] = rank
        row[f"passage_{i}_score"] = score
        row[f"passage_{i}_text"] = text

    row.update({
        "generated_answer": generated_answer,
        "predicted_answer_letter": predicted_letter,
        "is_correct": is_correct,
        "error_message": error_message or "",
    })

    # --- 3 champs d'annotation (alignés sur le Protocole B) ---
    row["retrieval_relevance_1_5"] = ""       # 1-5
    row["retrieval_sufficiency_0_2"] = ""     # 0-2
    # Primary Error Source : uniquement si is_correct = False ; options
    # via data validation Excel (cf. apply_data_validation).
    row["primary_error_source"] = "" if not is_correct else NOT_APPLICABLE_LABEL
    row["annotator_notes"] = ""

    return row


def apply_data_validation(path: str, max_row: int = 500):
    """Ajoute des listes déroulantes Excel pour les champs d'annotation."""
    if not os.path.exists(path):
        return
    wb = load_workbook(path)
    ws = wb.active
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]

    error_source_formula = '"' + ",".join(ERROR_SOURCE_OPTIONS + [NOT_APPLICABLE_LABEL]) + '"'
    validations = {
        "retrieval_relevance_1_5": '"1,2,3,4,5"',
        "retrieval_sufficiency_0_2": '"0,1,2"',
        "primary_error_source": error_source_formula,
    }
    for col_name, formula in validations.items():
        if col_name not in header:
            continue
        col_idx = header.index(col_name) + 1
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        dv = DataValidation(type="list", formula1=formula, allow_blank=True, showDropDown=False)
        ws.add_data_validation(dv)
        dv.add(f"{col_letter}2:{col_letter}{max_row}")

    wb.save(path)
    wb.close()


def run_batch(llm_gen: HFGenerator, campaign_key: str, campaign_label: str,
              scenario_type: str, scenario_name: str, llm_label: str,
              sampled_qids: list, qcm_data: dict, retriever: EmbeddingRetriever,
              reranker_models: dict):
    out_paths = {a: campaign_batch_path(campaign_key, scenario_type, llm_label, scenario_name, a)
                 for a in ANNOTATORS}
    completed_per_annotator = {a: load_completed_keys_xlsx(p, OUTPUT_KEY_COLS)
                                for a, p in out_paths.items()}
    remaining = [
        qid for qid in sampled_qids
        if not all((llm_label, scenario_type, qid) in completed_per_annotator[a] for a in ANNOTATORS)
    ]

    if not remaining:
        print(f"⏭️  Campagne {campaign_key} / Scénario {scenario_type} ({scenario_name}) — {llm_label}: "
              f"déjà complet, skip.")
        return

    print(f"\n{'=' * 70}")
    print(f"📝 Campagne {campaign_key} ({campaign_label}) — Scénario {scenario_type} ({scenario_name}) — "
          f"{llm_label} [{len(remaining)}/{len(sampled_qids)} questions restantes]")
    print(f"{'=' * 70}")

    for i, qid in enumerate(remaining, 1):
        q_entry = qcm_data[qid]
        question = q_entry["question"]
        options = q_entry["options"]
        options_text = "\n".join(f"{k}: {v}" for k, v in options.items())
        correct_letter = q_entry["answer_letter"]

        try:
            t0 = time.time()
            passages, _pool = retrieve_with_logging(retriever, scenario_name, reranker_models, question)
            context = "\n".join(text for _, _, text in passages)
            prompt = config.FIQH_TEMPLATE.format(
                context=context, question=question, options=options_text,
                letters=config.letters_comma(options),
            )
            model_answer = llm_gen.invoke(prompt)
            gen_time = time.time() - t0

            is_correct = validate_answer(model_answer, correct_letter)
            predicted_letter = extract_answer_letter(model_answer)

            for a in ANNOTATORS:
                if (llm_label, scenario_type, qid) in completed_per_annotator[a]:
                    continue
                row = build_row(
                    campaign_key, campaign_label, scenario_type, scenario_name,
                    llm_label, FIXED_EMBEDDING, qid, question, options_text,
                    correct_letter, passages, model_answer, is_correct, predicted_letter,
                    annotator=a,
                )
                append_row_xlsx(out_paths[a], row)

            status = "✅" if is_correct else "❌"
            print(f"   [{i}/{len(remaining)}] Q{qid} {status} gen={gen_time:.2f}s")

        except Exception as e:
            print(f"   ⚠️ Erreur Q{qid}: {e}")
            for a in ANNOTATORS:
                if (llm_label, scenario_type, qid) in completed_per_annotator[a]:
                    continue
                error_row = build_row(
                    campaign_key, campaign_label, scenario_type, scenario_name,
                    llm_label, FIXED_EMBEDDING, qid, question, options_text,
                    correct_letter, [], "", "", "", annotator=a, error_message=str(e),
                )
                append_row_xlsx(out_paths[a], error_row)

    for a in ANNOTATORS:
        apply_data_validation(out_paths[a])


def run_campaign(campaign_key: str, qcm_data: dict, all_question_ids: list,
                  device: str, quantize_4bit: bool, seed: int):
    spec = CAMPAIGN_SPECS[campaign_key]
    llm_label = spec["llm"]
    embedding_name = spec["embedding"]
    scenario_a = spec["scenario_a"]
    scenario_b = spec["scenario_b"]
    campaign_label = spec["label"]

    rag_path = rag_results_path(llm_label)
    rag_all_results = load_all_results_deduped(rag_path, qcm_data=qcm_data)
    map_a = rag_all_results.get((embedding_name, scenario_a), {})
    map_b = rag_all_results.get((embedding_name, scenario_b), {})

    if not map_a:
        print(f"❌ {llm_label}: résultats introuvables pour {embedding_name}+{scenario_a} dans {rag_path}.")
        print(f"   -> python run_generation.py --llm {llm_label} --mode rag --embedding {embedding_name} "
              f"--scenario \"{scenario_a}\"")
        return
    if not map_b:
        print(f"❌ {llm_label}: résultats introuvables pour {embedding_name}+{scenario_b} dans {rag_path}.")
        print(f"   -> python run_generation.py --llm {llm_label} --mode rag --embedding {embedding_name} "
              f"--scenario \"{scenario_b}\"")
        return
    if len(map_a) < len(all_question_ids) or len(map_b) < len(all_question_ids):
        print(f"❌ {llm_label}: résultats incomplets (A: {len(map_a)}/{len(all_question_ids)}, "
              f"B: {len(map_b)}/{len(all_question_ids)}). Complétez avant de relancer.")
        return

    basis_map = map_b if SAMPLING_BASIS == "B" else map_a
    sampled_qids = sample_fixed_20_20(all_question_ids, basis_map, seed)
    n_incorrect_basis = sum(1 for qid in sampled_qids if not basis_map[qid])
    print(f"   📊 Campagne {campaign_key} ({llm_label}): échantillon de {len(sampled_qids)} questions, "
          f"tiré sur le Scénario {SAMPLING_BASIS} ({n_incorrect_basis} incorrectes / "
          f"{len(sampled_qids) - n_incorrect_basis} correctes selon ce scénario)")
    n_a_correct = sum(1 for qid in sampled_qids if map_a[qid])
    n_b_correct = sum(1 for qid in sampled_qids if map_b[qid])
    print(f"   ℹ️  Dans ce même échantillon : Scénario A ({scenario_a}) = {n_a_correct}/{len(sampled_qids)} "
          f"correctes ; Scénario B ({scenario_b}) = {n_b_correct}/{len(sampled_qids)} correctes")

    needed_reranker_names = set()
    for scenario_name in (scenario_a, scenario_b):
        for reranker_name, _ in config.RERANKING_SCENARIOS_DEF[scenario_name]["stages"]:
            needed_reranker_names.add(reranker_name)

    reranker_models = {}
    if needed_reranker_names:
        all_rerankers = load_all_rerankers(device)
        reranker_models = {k: v for k, v in all_rerankers.items() if k in needed_reranker_names}

    print(f"\n{'#' * 70}\n# Campagne {campaign_key}: {campaign_label} — LLM: {llm_label}\n{'#' * 70}")
    llm_gen = None
    retriever = None
    try:
        llm_gen = HFGenerator(config.LLM_MODEL_IDS[llm_label], device=device, quantize_4bit=quantize_4bit)
        embedding_path = config.EMBEDDING_MODEL_PATHS[embedding_name]
        retriever = EmbeddingRetriever(embedding_path, device=device)

        run_batch(llm_gen, campaign_key, campaign_label, "A", scenario_a, llm_label,
                  sampled_qids, qcm_data, retriever, reranker_models)
        run_batch(llm_gen, campaign_key, campaign_label, "B", scenario_b, llm_label,
                  sampled_qids, qcm_data, retriever, reranker_models)

    except Exception as e:
        print(f"❌ Échec pour la campagne {campaign_key}: {e}")
        import traceback
        traceback.print_exc()
    finally:
        if retriever is not None:
            retriever.unload()
        if llm_gen is not None:
            llm_gen.unload()
        if reranker_models:
            unload_rerankers(reranker_models)


def main():
    parser = argparse.ArgumentParser(
        description="Construit les lots d'annotation des Campagnes 2 (Fanar-9B, CAMeLBERT, "
                     "mini_reranker vs cross_encoder) et 3 (Llama3-8B, CAMeLBERT, dual_mini_ara vs "
                     "dual_mini_cross), échantillon fixe 20 correctes / 20 incorrectes, 3 champs "
                     "d'annotation (Relevance, Sufficiency, Primary Error Source)."
    )
    parser.add_argument("--campaign", required=True, choices=["2", "3", "all"])
    parser.add_argument("--seed", type=int, default=SEED_DEFAULT)
    parser.add_argument("--device", choices=["cpu", "cuda"], default=config.DEFAULT_DEVICE)
    parser.add_argument("--no-4bit", action="store_true")
    args = parser.parse_args()

    if not os.path.exists(config.QCM_FILE):
        print(f"❌ Fichier non trouvé: {config.QCM_FILE}")
        return
    with open(config.QCM_FILE, "r", encoding="utf-8") as f:
        qcm_data = json.load(f)
    all_question_ids = list(qcm_data.keys())

    campaigns_to_run = ["2", "3"] if args.campaign == "all" else [args.campaign]
    quantize_4bit = not args.no_4bit

    print_environment_banner(args.device)

    for campaign_key in campaigns_to_run:
        run_campaign(campaign_key, qcm_data, all_question_ids, args.device, quantize_4bit, args.seed)

    print("\n✅ Construction des lots Campagnes 2/3 terminée (ou interrompue proprement — "
          "relancez la même commande pour reprendre).")
    print(f"   Fichiers de sortie (.xlsx) dans: {ANNOTATION_RESULTS_DIR}/")


if __name__ == "__main__":
    main()
